import openpyxl
import pandas as pd
import datetime
print(datetime.datetime.now())

wbTitles = openpyxl.load_workbook("Path_to_titles_file")
wsTitles = wbTitles['Name_of_sheet']

  
list_of_input_files = []
wb_list_of_input_files = openpyxl.load_workbook("Paths_file")
ws_list_of_input_files = wb_list_of_input_files["Name_of_sheet"]


for r in range(1, ws_list_of_input_files.max_row + 1):
    list_of_input_files.append(ws_list_of_input_files.cell(row = r, column = 1).value)
    
    
    
for input_file in list_of_input_files: # loop which suppling output_files from input_file
    try:  # check path correctness for each output_file  
        wb_input_file = openpyxl.load_workbook(input_file)
        ws_input_file = wb_input_file['Name_of_sheet']
    except FileNotFoundError as e:
        print("Path: %s is not correct" % input_file)
        continue
        

    id_start_from = 0  # define from which row list of id are
    
    for i in range(1,ws_input_file.max_row):
        if ws_input_file.cell(row = i , column = 1).value == "ID" and ws_input_file.cell(row = i + 1 , column = 1).value == "tekst" \
        and  type(ws_input_file.cell(row = i + 2 , column = 1).value) == int:
            id_start_from = i + 2
            
        elif ws_input_file.cell(row = i , column = 1).value == "ID" and type(ws_input_file.cell(row = i + 1 , column = 1).value) == int:
            id_start_from = i + 1
            
            
            
    start_adding_id = 1  # determines from which row data are pasted
    while wsTitles.cell(row = start_adding_id, column = 7).value != None and wsTitles.cell(row = start_adding_id, column = 9).value != None:
        start_adding_id += 1
    

    for r in range(id_start_from, ws_input_file.max_row):  #adding data from each input_file to sum_of_files
        for c in range(1, 35):
            wsTitles.cell(row = start_adding_id + r - id_start_from, column = c).value = ws_input_file.cell(row = r , column = c).value
            

wbSum_of_files = wbTitles
wbSum_of_files.save("Path_to_file_sum") 
print(datetime.datetime.now())