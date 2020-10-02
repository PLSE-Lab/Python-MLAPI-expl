import openpyxl
import pandas as pd
import datetime
print(datetime.datetime.now())


wb_configurate_file = openpyxl.load_workbook("Path_to_configurate_file")
ws_configurate_file = wb_configurate_file["Name of sheet"]

list_of_building_id = []
wb_input_file = openpyxl.load_workbook(ws_configurate_file.cell(row = 3 , column = 2).value)
ws_input_file = wb_input_file["Name of sheet"]
input_file_Pd = pd.read_excel(ws_configurate_file.cell(row = 3 , column = 2).value)


for r in range(2, ws_input_file.max_row + 1): # creating id list from input file
    list_of_building_id.append(ws_input_file.cell(row = r, column = 2).value)
    
    
    
list_of_output_files = []
wb_list_of_output_files = openpyxl.load_workbook(ws_configurate_file.cell(row = 8 , column = 2).value)
ws_list_of_output_files = wb_list_of_output_files["Name of sheet"]


for r in range(1, ws_list_of_output_files.max_row + 1): # creat list_of_output_files from file with outputs
    list_of_output_files.append(ws_list_of_output_files.cell(row = r, column = 1).value)
    
    
for output_file in list_of_output_files: # loop which suppling output_files from input_file
    try:  # check path correctness for each output_file  
        file = pd.read_excel(output_file)
    except FileNotFoundError as e:
        print("Path: %s is not correct" % output_file)
        continue

    
    try: # Chcecking available to save for each output_file 
        file = openpyxl.load_workbook(output_file)
        file.save(output_file)
    except PermissionError as per:
        print("The file with the path %s is opened by another user" % output_file)
        continue  
    
    if output_file[output_file.rfind("."):] == ".xlsm":  # for .xlsm files keep macros
        wb_output_file = openpyxl.load_workbook(output_file, read_only=False, keep_vba=True)
    else:
        wb_output_file = openpyxl.load_workbook(output_file)
        
    ws_output_file = wb_output_file['Name of sheet']
    
    
    id_start_from = 0  # define from which row list of id are
    for i in range(1,ws_output_file.max_row):
        if ws_output_file.cell(row = i , column = 1).value == "Id_of_index" and ws_output_file.cell(row = i + 1 , column = 1).value == "tekst" \
        and  type(ws_output_file.cell(row = i + 2 , column = 1).value) == int:
            id_start_from = i + 2
            date_actualization_row = i - 2
        elif ws_output_file.cell(row = i , column = 1).value == "Id_of_index" and type(ws_output_file.cell(row = i + 1 , column = 1).value) == int:
            id_start_from = i + 1
            date_actualization_row = i - 2
        

    id_output_file = []
    
    # creat list_of_id_output_file from each output_file
    for r in range(id_start_from, ws_output_file.max_row):
        id_output_file.append(ws_output_file.cell(row = r, column = 1).value)
    
    # list of columns from input_file
    columns = input_file_Pd[input_file_Pd["Id_of_index"].isin(list(set(list_of_building_id).intersection(id_output_file))) ]
    
    # columns_to_clean before supplementing with new data 
    columns_to_clean = ws_configurate_file.cell(row = 13 , column = 2).value.split(",")
    
    # cleaning algoritm
    for column in list(set(columns_to_clean).intersection(list(columns["Column"].unique()))):
        for i in range(id_start_from, ws_output_file.max_row):
            ws_output_file.cell(row = i, column = column).value = ""
            
    # supplementing output_file specific cell based on the id from input_file and output_file (row definition) 
    # column is defined by the column number from input_file
    for building in list(set(list_of_building_id).intersection(id_output_file)):
        for i in range(1, ws_output_file.max_row):
            if ws_output_file.cell(row = i, column = 1).value == building:
                for j in range(1, len(list_of_building_id)):
                    if ws_input_file.cell(row = j, column = 2).value == ws_output_file.cell(row = i, column = 1).value:
                        ws_output_file.cell(row = i, column = ws_input_file.cell(row = j, column = 1).value ).value = ws_input_file.cell(row = j, column = 3).value
    
    
    # adding date of supplementing each column
    for i in list(columns["Column"].unique()):
        ws_output_file.cell(row = date_actualization_row, column = i).value = str(datetime.date.today())
    
    wb_output_file.save(output_file)
    print(datetime.datetime.now())   
print(datetime.datetime.now())