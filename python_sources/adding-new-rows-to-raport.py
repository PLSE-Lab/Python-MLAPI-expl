import openpyxl
import pandas as pd
import datetime
print(datetime.datetime.now())


# wsadowy
wb_input_file = openpyxl.load_workbook("Path_to_input_file")
ws_input_file = wb_input_file["Name_of_sheet"]
input_file_Pd = pd.read_excel("Path_to_input_file")

list_of_output_files = []
wb_list_of_output_files = openpyxl.load_workbook("Paths_file")
ws_list_of_output_files = wb_list_of_output_files["Name_of_sheet"]




for r in range(1, ws_list_of_output_files.max_row + 1): # creat list_of_output_files from file with outputs
    list_of_output_files.append(ws_list_of_output_files.cell(row = r, column = 1).value)
    

for output_file in list_of_output_files: # loop which suppling output_files from input_file
    try:  # check path correctness for each output_file  
        file = pd.read_excel(output_file)
    except FileNotFoundError as e:
        print("Path: %s is not correct" % output_file)
        continue

    
    try: # Chcecking available to save for each output_file 
        file = openpyxl.load_workbook(output_file)
        file.save(output_file)
    except PermissionError as per:
        print("The file with the path %s is opened by another user" % output_file)
        continue  
    
    if output_file[output_file.rfind("."):] == ".xlsm":  # for .xlsm files keep macros
        wb_output_file = openpyxl.load_workbook(output_file, read_only=False, keep_vba=True)
    else:
        wb_output_file = openpyxl.load_workbook(output_file)
        
    ws_output_file = wb_output_file['Name_of_sheet']
        

    id_start_from = 0  # define from which row list of id are
    for i in range(1,ws_output_file.max_row):
        if ws_output_file.cell(row = i , column = 1).value == "Id_of_index" and ws_output_file.cell(row = i + 1 , column = 1).value == "tekst" \
        and  type(ws_output_file.cell(row = i + 2 , column = 1).value) == int:
            id_start_from = i + 2
            date_actualization_row = i - 2
        elif ws_output_file.cell(row = i , column = 1).value == "Id_of_index" and type(ws_output_file.cell(row = i + 1 , column = 1).value) == int:
            id_start_from = i + 1
            date_actualization_row = i - 2
    

    id_output_file = []

    for r in range(id_start_from, ws_output_file.max_row):
        id_output_file.append(ws_output_file.cell(row = r, column = 1).value)

        
    list_of_id_per_file = []
    
    [list_of_id_per_file.append(ws_input_file.cell(row = r, column = 1).value) for r in range(2, ws_input_file.max_row + 1) if ws_input_file.cell(row = r, column = 2).value == ws_output_file.cell(row = id_start_from , column = 7).value and ws_input_file.cell(row = r, column = 3).value == ws_output_file.cell(row = id_start_from , column = 6).value]
    
 
    
    def diff(first, second):
        second = set(second)
        return [item for item in first if item not in second]
    
    print(diff(list_of_id_per_file, id_output_file))
    print(output_file)
    for i in range(0,len(diff(list_of_id_per_file, id_output_file))):
        ws_output_file.cell(row = ws_output_file.max_row + i , column = 1).value = diff(list_of_id_per_file, id_output_file)[i]
    
    wb_output_file.save(output_file)
    print(datetime.datetime.now(), end="\n\n")   
print(datetime.datetime.now())