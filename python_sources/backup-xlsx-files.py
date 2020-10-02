# backup all files which paths are in path_file to backup

import openpyxl
import pandas as pd

list_of_files = [] 
wblist_of_files = openpyxl.load_workbook("path_file")
wslist_of_files = wblist_of_files["Name of sheet"]

for r in range(1, wslist_of_files.max_row + 1):
    list_of_files.append(wslist_of_files.cell(row = r, column = 1).value)
    
for path_to_file in list_of_files:
    file = openpyxl.load_workbook(path_to_file)
    backup = "path_to_backup_file" + path_to_file[path_to_file.rfind("\\")+1:]
    file.save(backup)